import openpyxl
import os
import pandas as pd
from operator import le
import datetime
import mysql.connector
mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="123456",
  database="mydata"
)
mycursor = mydb.cursor()

class Person:
    def __init__(self,Id,firstName,father,lastName,Gender,BirthYear,Address):
            if Id>0 and len(firstName)>1 and len(father)>1 and len(lastName)>1:
                self._Id=Id
                self._firstName=firstName
                self._father=father
                self._lastName=lastName
                self._Gender=Gender
                self._BirthYear=BirthYear
                self._Address=Address
            else:
                raise Exception("Invalid Data")
    #start id
    def setId(self,newId):
        if newId>0:
            self._Id=newId
        else:
            raise Exception("Invalid Id")
    def getId(self):
        return self._Id
    #end id
    #start fn
    def setFn(self,newFn):
        if len(newFn)>1:
            self._firstName=newFn
        else:
            raise Exception("Invalid FirstName")
    def getFn(self):
        return self._firstName
    #end fn
    #start fa
    def setFa(self,newFa):
        if len(newFa)>1:
            self._father=newFa
        else:
            raise Exception("Invalid Father")
    def getFa(self):
        return self._father
    #end fa
    #start ln
    def setln(self,newln):
        if len(newln)>1:
            self._lastName=newln
        else:
            raise Exception("Invalid LastName")
    def getLn(self):
        return self._firstName
    #end ln
    #start Gender
    def setGn(self,newGn):
        if newGn=="male" or newGn=="female":
            self._Gender=newGn
        else:
            raise Exception("Invalid Gender")
    def getGn(self):
        return self._Gender
    #end Gender
    # #start By
    # def setBy(self,newBy):
    #     if newBy>1900:
    #    self._BirthYear=newBy
    #     else:
    #         raise Exception("Invalid Birthyear")
    def getBy(self):
        return self._BirthYear
    # #end By
    
    #start address
    def setAd(self,newAd):
        if len(newAd)>1:
            self._Address=newAd
        else:
            raise Exception("Invalid Address")
    def getAd(self):
        return self._Address
    #end address
class Visitings:#Person مركية مع 
    def __init__(self,Id,DateVisited,PersonId,VisitorName,MountinMinutes):
            if Id>0 and PersonId>0 and len(VisitorName)>1 :
                self._Id=Id
                self._DateVisited=DateVisited
                self._PersonId=PersonId
                self._VisitorName=VisitorName
                self._MountinMinutes=MountinMinutes
            else:
                raise Exception("Invalid Data")
    #start id
    def setId(self,newId):
        if newId>0:
            self._Id=newId
        else:
            raise Exception("Invalid Id")
    def getId(self):
        return self._Id
    #end id
    #start Vn
    def setVn(self,newVn):
        if len(newVn)>1:
            self._VisitorName=newVn
        else:
            raise Exception("Invalid FirstName")
    def getVn(self):
        return self._VisitorName
    #end Vn
    #start DT
    def setDt(self,newDt):
        self._DateVisited=newDt
    def getDt(self):
        return self._DateVisited
    #end DT
    #start PID
    def setPId(self,newPId):
        if newPId>0:
            self._PersonId=newPId
        else:
            raise Exception("Invalid Id")
    def getPId(self):
        return self._PersonId
    #end PID
    #start MoM
    def setMom(self,newMom):
        if newMom>0 and newMom<30:
            self._MountinMinutes=newMom
        else:
            raise Exception("Invalid Id")
    def getMom(self):
        return self._MountinMinutes
    #end MoM
    
class Convicts:#Person مركبة مع 
    def __init__(self,Id,fromDate,toDate,personId,offenseId):
        if Id>0 and personId>0 and offenseId>0:     
            self._Id=Id
            self._fromDate=fromDate
            self._toDate=toDate
            self._personId=personId
            self._offenseId=offenseId
        else:
            raise Exception("Invalid Data")
    #start id
    def setId(self,newId):
        if newId>0:
            self._Id=newId
        else:
            raise Exception("Invalid Id")
    def getId(self):
        return self._Id
    #end id
    #start fromDate
    def setFd(self,newFd):
        #if len(newFd)>4:
        self._fromDate=newFd
        #else:
         #   raise Exception("Invalid From Date")
    def getFd(self):
        return self._fromDate
    #end fromDate
    
    #start toDate
    def setTd(self,newTd):
        #if len(newTd)>4:
        self._toDate=newTd
        #else:
        #    raise Exception("Invalid To Date")
    def getTd(self):
        return self._toDate
    #end toDate
    #start Pid
    def setPId(self,newPId):
        if newPId>0:
            self._personId=newPId
        else:
            raise Exception("Invalid PersonId")
    def getPId(self):
        return self._personId
    #end Pid
    #start Oid
    def setOId(self,newOId):
        if newOId>0:
            self._personId=newOId
        else:
            raise Exception("Invalid Offence Id")
    def getOId(self):
        return self._offenseId
    #end Oid
    
class Offense:#Convicts مركبة مع 
    def __init__(self,Id,name):
        if Id>0 and len(name)>1:
            self._Id=Id
            self._name=name
        else:
            raise Exception("Invalid Data")
    #start id
    def setId(self,newId):
        if newId>0:
            self._Id=newId
        else:
            raise Exception("Invalid Id")
    def getId(self):
        return self._Id
    #end id
    #start fn
    def setOn(self,newOn):
        if len(newOn)>1:
            self._name=newOn
        else:
            raise Exception("Invalid OffencName")
    def getOn(self):
        return self._name
    #end fn
    
class DungeonMoves:
    def __init__(self,Id,dungeonId,personId,fromDate):
        if Id>0 and dungeonId>0 and personId>0:
            self._Id=Id
            self._dungeonId=dungeonId#Dungeon
            self._personId=personId #Person
            self._fromDate=fromDate
        else:
            raise Exception("Invalid Data")
    #start id
    def setId(self,newId):
        if newId>0:
            self._Id=newId
        else:
            raise Exception("Invalid Id")
    def getId(self):
        return self._dungeonId
    #end id
    #start Duid
    def setDuId(self,newDuId):
        if newDuId>0:
            self._dungeonId=newDuId
        else:
            raise Exception("Invalid dungeon Id")
    def getDuId(self):
        return self._dungeonId
    #end Duid
    #start Pid
    def setPId(self,newPId):
        if newPId>0:
            self._personId=newPId
        else:
            raise Exception("Invalid PersonId")
    def getPId(self):
        return self._personId
    #end Pid
    #start fromDate
    def setFd(self,newFd):
        if len(newFd)>4:
            self._fromDate=newFd
        else:
            raise Exception("Invalid From Date")
    def getFd(self):
        return self._fromDate
    #end fromDate
class Dungeon:
    def __init__(self,Id,name,size):
        if Id>0 and len(name)>1:
            self._Id=Id
            self._name=name
            self._size=size
        else:
            raise Exception("Invalid Data")
    #start id
    def setId(self,newId):
        if newId>0:
            self._Id=newId
        else:
            raise Exception("Invalid Id")
    def getId(self):
        return self._Id
    #end id
    #start Vn
    def setNa(self,newNa):
        if len(newNa)>1:
            self._name=newNa
        else:
            raise Exception("Invalid Name")
    def getNa(self):
        return self._name
    #end Vn
    #start size
    def setSize(self,newSize):
        if newSize>0:
            self._size=newSize
        else:
            raise Exception("Invalid Size")
    def getSize(self):
        return self._size
    #end size


def inserConvicts():
    isStop="-1"
    lst=[]
    while True:
        id=int(input("Id: "))
        fd=input("From Date: ")
        td=input("To Date: ")
        pId=int(input("PersonId: "))
        oId=int(input("OffenceId: "))
        p=Convicts(id,fd,td,pId,oId)
        for i in lst:
            i.append(p)
        sql="INSERT INTO convicts(Id,FromDate,ToDate,personId,offenceId) Values(%s,%s,%s,%s,%s)"
        val=(p.getId(),p.getFd(),p.getTd(),p.getPId(),p.getOId())
        mycursor.execute(sql,val)
        
        mydb.commit()
        #print("1 record inserted, ID:", mycursor.lastrowid)
      
        isStop=input("if you need to stop press 0: ")
        if isStop=="0":
            break

def showConvicts():#Show Convicts
    mycursor.execute("SELECT * FROM convicts")
    myresult = mycursor.fetchall()

    for x in myresult:
        print(x)
#s1=showConvicts()
#c1=inserConvicts()
def printToExcelConvict():
    sql = "select * from convicts;"
    df=pd.read_sql(sql,mydb)

    df.head()
    df.to_excel("convicts.xlsx",index=False)


def deleteFileConvict():
    import os
    if os.path.exists("convicts.xlsx"):
        os.remove("convicts.xlsx")
    else:
        print("The file does not exist")
dd=deleteFileConvict()
def readFileConvicts():
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook("convicts.xlsx")

    # Define variable to read the active sheet:
    worksheet = wookbook.active

    # Iterate the loop to read the cell values
    for i in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, worksheet.max_column):
            print(col[i].value, end="\t\t")
        print('')
def insertDirPer(id,fn,fa,ln,ge,by,ad):
    sql="INSERT INTO persons(Id,FirstName,Father,LastName,Gender,BirthYear,Address) Values(%s,%s,%s,%s,%s,%s,%s)"
    val=(id,fn,fa,ln,ge,by,ad)
    mycursor.execute(sql,val)
        
    mydb.commit()
#inp=insertDirPer(22,"hani","ahmad","ahmad","male","1999,1,1","uuii")

def insertPrison():
    isStop="-1"
    lst=[]
    while True:
        id=int(input("Id: "))
        fn=input("FirstName: ")
        fa=input("Father: ")
        ln=input("LastName: ")
        gn=input("Gender: ")
        by=input("BirthYear: ")
        ad=input("Address: ")
        p=Person(id,fn,fa,ln,gn,by,ad)
        for i in lst:   
            i.append(p)
        sql="INSERT INTO persons(Id,FirstName,Father,LastName,Gender,BirthYear,Address) Values(%s,%s,%s,%s,%s,%s,%s)"
        val=(p.getId(),p.getFn(),p.getFa(),p.getLn(),p.getGn(),p.getBy(),p.getAd())
        mycursor.execute(sql,val)
        
        mydb.commit()
        #print("1 record inserted, ID:", mycursor.lastrowid)
      
        isStop=input("if you need to stop press 0: ")
        if isStop=="0":
            break    
#q1=insertPrison()
def showprisoners():#Show All Prisoners
    mycursor.execute("SELECT * FROM persons")
    myresult = mycursor.fetchall()

    for x in myresult:
        print(x)

def printToExcelPrisoners():    
    sql = "select * from persons;"
    df=pd.read_sql(sql,mydb)

    df.head()
    df.to_excel("prioner.xlsx",index=False)
#pyii=printToExcelPrisoners()
def deleteFilePrison():
    import os
    if os.path.exists("prioner.xlsx"):
        os.remove("prioner.xlsx")
    else:
        print("The file does not exist")
def readFilePrison():
    if os.path.exists("prioner.xlsx"):
        # Define variable to load the wookbook
        wookbook = openpyxl.load_workbook("prioner.xlsx")

        # Define variable to read the active sheet:
        worksheet = wookbook.active

        # Iterate the loop to read the cell values
        for i in range(0, worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column):
                print(col[i].value, end="\t\t")
            print('')
    else:
        print("The File Not Found")

def insertOffenc():
    isStop="-1"
    lst=[]
    while True:
        id=int(input("Id: "))
        on=input("Offence Name: ")
        p=Offense(id,on)
        for i in lst:
            i.append(p)
        sql="INSERT INTO offence(Id,namee) Values(%s,%s)"
        val=(p.getId(),p.getOn())
        mycursor.execute(sql,val)
        
        mydb.commit()
        #print("1 record inserted, ID:", mycursor.lastrowid)
      
        isStop=input("if you need to stop press 0: ")
        if isStop=="0":
            break    
def showOffence():#Show All The Offence
    mycursor.execute("SELECT * FROM offence")
    myresult = mycursor.fetchall()

    for x in myresult:
        print(x)
def printToExcelOffence():
    sql = "select * from offence;"
    df=pd.read_sql(sql,mydb)

    df.head()
    df.to_excel("offence.xlsx",index=False)
def deleteFileOffence():
    import os
    if os.path.exists("offence.xlsx"):
        os.remove("offence.xlsx")
    else:
        print("The file does not exist")
def readFileOffene():
    if os.path.exists("offence.xlsx"):
        # Define variable to load the wookbook
        wookbook = openpyxl.load_workbook("offence.xlsx")

        # Define variable to read the active sheet:
        worksheet = wookbook.active

        # Iterate the loop to read the cell values
        for i in range(0, worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column):
                print(col[i].value, end="\t\t")
            print('')
    else:
        print("The File Not Found")


def insertVisitings():
    isStop="-1"
    lst=[]
    while True:
        id=int(input("Id: "))
        dv=input("DateVisited: ")
        pId=int(input("PersonId: "))
        vn=input("Visitor Name: ")
        mom=int(input("Mount of minutes: "))
        p=Visitings(id,dv,pId,vn,mom)
        for i in lst:
            i.append(p)
        sql="INSERT INTO Visitings(Id,DateVisited,personId,visitorName,MountMinutes) Values(%s,%s,%s,%s,%s)"
        val=(p.getId(),p.getDt(),p.getPId(),p.getVn(),p.getMom())
        mycursor.execute(sql,val)
        
        mydb.commit()
        #print("1 record inserted, ID:", mycursor.lastrowid)
      
        isStop=input("if you need to stop press 0: ")
        if isStop=="0":
            break
def showVisitings():#Show Visitors
    mycursor.execute("SELECT * FROM visitings")
    myresult = mycursor.fetchall()

    for x in myresult:
        print(x)
    
def printToExcelVisitings():
    sql = "select * from visitings;"
    df=pd.read_sql(sql,mydb)

    df.head()
    df.to_excel("Visitors.xlsx",index=False)
def deleteFileVisit():
    import os
    if os.path.exists("Visitors.xlsx"):
        os.remove("Visitors.xlsx")
    else:
        print("The file does not exist")
def readFileVisit():
    if os.path.exists("Visitors.xlsx"):
        # Define variable to load the wookbook
        wookbook = openpyxl.load_workbook("Visitors.xlsx")

        # Define variable to read the active sheet:
        worksheet = wookbook.active

        # Iterate the loop to read the cell values
        for i in range(0, worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column):
                print(col[i].value, end="\t\t")
            print('')
    else:
        print("The File Not Found")


def insertDungeon():
    isStop="-1"
    lst=[]
    while True:
        id=int(input("Id: "))
        na=input("Name: ")
        siz=int(input("Size m^2 : "))
        
        p=Dungeon(id,siz)
        for i in lst:
            i.append(p)
        sql="INSERT INTO dungeon(Id,name,size) Values(%s,%s,%s)"
        val=(p.getId(),p.getNa(),p.getSize())
        mycursor.execute(sql,val)
        
        mydb.commit()
        #print("1 record inserted, ID:", mycursor.lastrowid)
      
        isStop=input("if you need to stop press 0: ")
        if isStop=="0":
            break
def showDungeon():#show Dungeon
    mycursor.execute("SELECT * FROM dungeon")
    myresult = mycursor.fetchall()

    for x in myresult:
        print(x)
def printToExcelVisitings():
    sql = "select * from dungeon;"
    df=pd.read_sql(sql,mydb)

    df.head()
    df.to_excel("Dungeon.xlsx",index=False)
def deleteFileDun():
    import os
    if os.path.exists("Dungeon.xlsx"):
        os.remove("Dungeon.xlsx")
    else:
        print("The file does not exist")
def readFileDun():
    if os.path.exists("Dungeon.xlsx"):
        # Define variable to load the wookbook
        wookbook = openpyxl.load_workbook("Dungeon.xlsx")

        # Define variable to read the active sheet:
        worksheet = wookbook.active

        # Iterate the loop to read the cell values
        for i in range(0, worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column):
                print(col[i].value, end="\t\t")
            print('')
    else:
        print("The File Not Found")


def insertDungeonMoves():
    isStop="-1"
    lst=[]
    while True:
        id=int(input("Id: "))
        fd=input("From Date: ")
        dId=int(input("PersonId: "))
        pId=int(input("OffenceId: "))
        p=DungeonMoves(id,fd,dId,pId)
        for i in lst:
            i.append(p)
        sql="INSERT INTO dungeonmoves(Id,FromDate,dungeonId,personId) Values(%s,%s,%s,%s)"
        val=(p.getId(),p.getFd(),p.getDuId(),p.getPId())
        mycursor.execute(sql,val)
        
        mydb.commit()
        #print("1 record inserted, ID:", mycursor.lastrowid)
      
        isStop=input("if you need to stop press 0: ")
        if isStop=="0":
            break
def showDungeonMoves():#Show DungeonMoves
    mycursor.execute("SELECT * FROM dungeonmoves")
    myresult = mycursor.fetchall()

    for x in myresult:
        print(x)
def printToExcelDunMov():
    sql = "select * from dungeonmoves;"
    df=pd.read_sql(sql,mydb)

    df.head()
    df.to_excel("dungeonmoves.xlsx",index=False)
def deleteFileDubMo():
    import os
    if os.path.exists("dungeonmoves.xlsx"):
        os.remove("dungeonmoves.xlsx")
    else:
        print("The file does not exist")
def readFileDunMO():
    if os.path.exists("dungeonmoves.xlsx"):
        # Define variable to load the wookbook
        wookbook = openpyxl.load_workbook("dungeonmoves.xlsx")

        # Define variable to read the active sheet:
        worksheet = wookbook.active

        # Iterate the loop to read the cell values
        for i in range(0, worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column):
                print(col[i].value, end="\t\t")
            print('')
    else:
        print("The File Not Found")
d=deleteFilePrison()
s=printToExcelPrisoners()


#insertPrisons=insertPrison()
#showAllPrisoners=showprisoners()

#iis=insertOffenc()
#show=showOffence()

#inserstVisit=insertVisitings()
#show=showVisitings()

#insertCon=inserConvicts()
#showcon=showConvicts()




















# def insertPrison1():
#     isStop="-1"
#     lst=[]
#     while True:
#         id=int(input("Id: "))
#         fn=input("FirstName: ")
#         fa=input("Father: ")
#         ln=input("LastName: ")
#         #gn=input("Gender: ")
#         #by=int(input("BirthYear: "))
#         ad=input("Address: ")
#         sql="INSERT INTO persons(Id,FirstName,Father,LastName,Gender,BirthYear,Address) Values(%s,%s,%s,%s,%s,%s,%s)"
#         val=(id,fn,fa,ln,None,None,ad)
#         mycursor.execute(sql,val)
        
#         mydb.commit()
#         #print("1 record inserted, ID:", mycursor.lastrowid)
#         p=Person(id,fn,fa,ln,None,None,ad)
#         for i in lst:
#             i.append(p)
#         isStop=input("if you need to stop press 0: ")
#         if isStop=="0":
#             break    



# id=int(input("Id: "))
# fn=input("FirstName: ")
# fa=input("Father: ")
# ln=input("LastName: ")
# gn=input("Gender: ")
# #by=int(input("BirthYear: "))
# ad=input("Address: ")

# sql="INSERT INTO persons(Id,FirstName,Father,LastName,Gender,BirthYear,Address) Values(%s,%s,%s,%s,%s,%s,%s)"
# val=(id,fn,fa,ln,gn,2000/12/1,ad)
# mycursor.execute(sql,val)

# mydb.commit()
# print("1 record inserted, ID:", mycursor.lastrowid)


# id=int(input("Id: "))
# fn=input("FirstName: ")
# sql = "INSERT INTO persons (Id, FirstName) VALUES (%s, %s)"
# val = (id, fn)
# mycursor.execute(sql, val)

# mydb.commit()

# print("1 record inserted, ID:", mycursor.lastrowid)




# p1=Person(1, "ahmad", "majd", "dd", "male", 2000, "aleppo")
# p2=Person(2, "ramez", "majd", "dd", "male", 2000, "aleppo")
# p3=Person(3, "akram", "majd", "dd", "male", 2000, "aleppo")
# o1=Offense(1,"Theft")
# o2=Offense(2,"kill")
# c1=Convicts(1,2010,2020,1,1)
# c2=Convicts(2,2005,2050,2,2)
# lstC=[c1,c2]
# v1=Visitings(1, 2022, p1, "ahmad", "10m")
        
        